import math
import numpy as np
import pandas as pd
from sklearn.preprocessing import OneHotEncoder
from sklearn.model_selection import train_test_split
from scipy.stats import bernoulli
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import openpyxl
from openpyxl.utils import get_column_letter
import pickle as pkl

class Perceptron:
    """
    Clasa pentru antrenarea si testarea unei retele neuronale cu un singur strat ascuns

    Date membru:
    - train_x: datele de antrenare
    - test_x: datele de testare
    - train_y: etichetele de antrenare
    - test_y: etichetele de testare
    - W1, b1: ponderile si bias-urile pentru stratul ascuns
    - W2, b2: ponderile si bias-urile pentru stratul de iesire
    """
    def __init__(self, df):
        y = df['Race']
        X = df.drop(columns='Race')
        self.X = X
        self.train_x, self.test_x, self.train_y, self.test_y = train_test_split(X, y, test_size=0.1, random_state=42)
        self.train_x, self.test_x, self.train_y, self.test_y = np.array(self.train_x), np.array(self.test_x), np.array(self.train_y), np.array(self.test_y)
        self.__one_hot_encode__()
    
    def set_feature_names(self, feature_names):
        self.feature_names = feature_names

    def get_feature_names(self):
        return self.feature_names

    def __one_hot_encode__(self):
        encoder = OneHotEncoder(sparse_output=False)
        self.train_y = encoder.fit_transform(self.train_y.reshape(-1, 1))
        # self.test_y = encoder.fit_transform(self.test_y.reshape(-1, 1))

    def __weights_init_Xavier_Uniform__(self, nr_neuroni_strat_ascuns):
        low_bound = -math.sqrt(6 / (self.train_x.shape[1] + nr_neuroni_strat_ascuns)) 
        upper_bound = math.sqrt(6 / (self.train_x.shape[1] + nr_neuroni_strat_ascuns))

        self.W1 = np.random.uniform(low=low_bound, high=upper_bound, size=(self.train_x.shape[1], nr_neuroni_strat_ascuns))
        self.b1 = np.zeros((1, nr_neuroni_strat_ascuns))

        low_bound = -math.sqrt(6 / (nr_neuroni_strat_ascuns + self.train_y.shape[1])) 
        upper_bound = math.sqrt(6 / (nr_neuroni_strat_ascuns + self.train_y.shape[1]))

        self.W2 = np.random.uniform(low=low_bound, high=upper_bound, size=(nr_neuroni_strat_ascuns, self.train_y.shape[1]))  
        self.b2 = np.zeros((1, self.train_y.shape[1]))

    def __sigmoid__(self, Z):
        return 1/(1 + np.exp(-Z))

    def __softmax__(self, Z):
        expZ = np.exp(Z - np.max(Z, axis=1, keepdims=True))
        return expZ / np.sum(expZ, axis=1, keepdims=True)
    
    def __forward_pass__(self, X, W1, b1, W2, b2, dropout_rate=0):
        Z1 = np.dot(X, W1) + b1
        A1 = self.__sigmoid__(Z1) # functia de activare pentru stratul ascuns

        dropouts = bernoulli.rvs(1 - dropout_rate, size=A1.shape)
        A1 = A1 * dropouts / (1 - dropout_rate)

        Z2 = np.dot(A1, W2) + b2
        A2 = self.__softmax__(Z2) # functia de activare pentru stratul de iesire

        return A1, A2, dropouts
    
    def __backward_pass__(self, X, y, A1, A2, W2, dropouts=None, dropout_rate=0):
        m = y.shape[0]

        dZ2 = A2 - y  # Derivative of softmax + cross-entropy
        dW2 = np.dot(A1.T, dZ2) / m  # Gradient w.r.t. weights (output layer)
        db2 = np.sum(dZ2, axis=0, keepdims=True) / m  # Gradient w.r.t. biases (output layer)
    
        # Hidden layer error term (dZ1)
        dZ1 = np.dot(dZ2, W2.T) * A1 * (1 - A1)  # For sigmoid activation (derivative of sigmoid)

        dZ1 = dZ1 * dropouts
        dZ1 /= (1 - dropout_rate)

        dW1 = np.dot(X.T, dZ1) / m  # Gradient w.r.t. weights (hidden layer)
        db1 = np.sum(dZ1, axis=0, keepdims=True) / m  # Gradient w.r.t. biases (hidden layer)
    
        return dW1, db1, dW2, db2
    
    def __compute_loss_cross_entropy__(self, y, A2):
        m = y.shape[0]
        log_likelihood = -np.log(A2[range(m), y.argmax(axis=1)])
        loss = np.sum(log_likelihood) / m
        return loss
    
    def __update_weights__(self, W1, b1, W2, b2, dW1, db1, dW2, db2, learning_rate):
        updated_W1 = W1 - learning_rate * dW1
        updated_b1 = b1 - learning_rate * db1
        updated_W2 = W2 - learning_rate * dW2
        updated_b2 = b2 - learning_rate * db2
        return updated_W1, updated_b1, updated_W2, updated_b2
    
    def antreneaza(self, nr_neuroni_strat_ascuns=100, rata_de_invatare=0.01, epoci=150, batch_size=64):
        mean_losses = []
        self.__weights_init_Xavier_Uniform__(nr_neuroni_strat_ascuns)

        for epoca in range(epoci):
            perm = np.random.permutation(self.train_x.shape[0])
            X = self.train_x[perm]
            y = self.train_y[perm]

            for i in range (0, X.shape[0], batch_size):
                mean_loss_per_batch = 0
                X_batch = X[i:i+batch_size]
                y_batch = y[i:i+batch_size]

                # Forward pass
                A1, A2, dropouts = self.__forward_pass__(X_batch, self.W1, self.b1, self.W2, self.b2, dropout_rate=0.2)

                # Compute loss
                loss = self.__compute_loss_cross_entropy__(y_batch, A2)
                mean_loss_per_batch += loss

                # Backward pass
                dw1, db1, dw2, db2 = self.__backward_pass__(X_batch, y_batch, A1, A2, self.W2, dropouts, dropout_rate=0.2)

                # Update weights
                self.W1, self.b1, self.W2, self.b2 = self.__update_weights__(self.W1, self.b1, self.W2, self.b2, dw1, db1, dw2, db2, rata_de_invatare)

            print(f'Epoch {epoca+1}/{epoci}, Loss: {loss:.4f}')
            mean_loss_per_batch /= math.ceil(X.shape[0] / batch_size)
            mean_losses.append(mean_loss_per_batch)

        return self.W1, self.b1, self.W2, self.b2, mean_losses
    
    def predict(self, X, W1, b1, W2, b2):
        _, A2, _ = self.__forward_pass__(X, W1, b1, W2, b2)
        return A2
    
    def save_model(self, filename):
        with open(filename, 'wb') as f:
            pkl.dump(self, f)
        print(f'Model saved to {filename}')

    @staticmethod
    def load_model(filename):
        with open(filename, 'rb') as f:
            model = pkl.load(f)
        print(f'Model loaded from {filename}')
        return model
    
    def ploteaza_loss(self, mean_losses):
        plt.plot(mean_losses)
        plt.xlabel('Epochs')
        plt.ylabel('Loss')
        plt.show()
    
    def vizualizeaza_puncte_eronate(self, X, y, predictions, hide_instances=False):
        if predictions.ndim > 1:
            predictions = np.argmax(predictions, axis=1)
                    
        df_x = pd.DataFrame(X, columns=self.X.columns)
        df_y = pd.DataFrame(y, columns=['true_label'])
        df_pred = pd.DataFrame(predictions, columns=['predicted_label'])
        df = pd.concat([df_x, df_y, df_pred], axis=1)

        df.to_excel('puncte_eronate.xlsx', index=False)

        # Load the workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        # Open the Excel file
        wb = openpyxl.load_workbook('puncte_eronate.xlsx')
        ws = wb.active

        # Define the fill colors for correct and incorrect predictions
        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        # Apply conditional formatting to the rows
        for row in range(2, len(df) + 2):  # Start from row 2 to skip the header
            true_label = ws.cell(row=row, column=len(df_x.columns) + 1).value
            predicted_label = ws.cell(row=row, column=len(df_x.columns) + 2).value
            fill = green_fill if true_label == predicted_label else red_fill
            for col in range(1, len(df.columns) + 1):
                ws.cell(row=row, column=col).fill = fill
        
        if hide_instances: 
            start_col = get_column_letter(1)
            end_col = get_column_letter(len(df_x.columns))
            ws.column_dimensions.group(start_col, end_col, hidden=True)

        # Save the workbook
        wb.save('puncte_eronate.xlsx')